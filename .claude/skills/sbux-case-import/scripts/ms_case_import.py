#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MeterSphere 用例批量导入/更新
从 Excel 读取用例数据，构建 MS 请求，创建或更新用例并回写结果

支持两种模式：
- import: 导入新用例（筛选 status=待导入）
- update: 更新已有用例（通过 --case-ids 指定编号）
"""

import argparse
import base64
import json
import sys
import time
import uuid
from typing import Dict, List, Optional, Any

try:
    import requests
except ImportError:
    print("Error: requests module not found. Please install: pip3 install requests", file=sys.stderr)
    sys.exit(1)

try:
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives import padding
except ImportError:
    print("Error: cryptography module not found. Please install: pip3 install cryptography", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl module not found. Please install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# Excel 列映射 (0-indexed)
COL_API_ID = 0            # A: api_id
COL_API_UUID = 1          # B: api_uuid
COL_CASE_ID = 2           # C: case_id
COL_CASE_UUID = 3         # D: case_uuid
COL_STATUS = 4            # E: status
COL_NAME = 5              # F: name
COL_PRIORITY = 6          # G: priority
COL_TEST_SUMMARY = 7      # H: test_summary (仅参考，不导入)
COL_DESCRIPTION = 8       # I: description
COL_TAGS = 9              # J: tags
COL_API_PATH = 10         # K: api_path
COL_API_METHOD = 11       # L: api_method
COL_HEADERS = 12          # M: headers (JSON)
COL_BODY_TYPE = 13        # N: body_type
COL_BODY = 14             # O: body
COL_ASSERT_JSON = 15      # P: assert_json (JSON)
COL_ASSERT_REGEX = 16     # Q: assert_regex (JSON)
COL_ASSERT_DURATION = 17  # R: assert_duration
COL_PRE_SCRIPT = 18       # S: pre_script
COL_POST_JDBC = 19        # T: post_jdbc (JSON)
COL_POST_SCRIPT = 20      # U: post_script


def generate_signature(access_key: str, secret_key: str) -> str:
    """生成 API 签名（AES-128-CBC）"""
    timestamp = str(int(time.time() * 1000))
    plain_text = f"{access_key}|{timestamp}"

    key = secret_key.encode('utf-8')[:16].ljust(16, b'\x00')
    iv = access_key.encode('utf-8')[:16].ljust(16, b'\x00')

    padder = padding.PKCS7(128).padder()
    padded_data = padder.update(plain_text.encode('utf-8')) + padder.finalize()

    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted = encryptor.update(padded_data) + encryptor.finalize()

    return base64.b64encode(encrypted).decode('utf-8')


def parse_json_field(value: Any, field_name: str) -> Any:
    """解析 JSON 字段，返回解析后的对象或空列表"""
    if not value:
        return []
    if isinstance(value, (list, dict)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError as e:
        print(f"Warning: Failed to parse {field_name}: {e}", file=sys.stderr)
        return []


def build_hash_tree(row_data: Dict) -> List[Dict]:
    """根据行数据构建 hashTree 数组"""
    hash_tree = []

    # 1. 前置脚本 (S列)
    pre_script = row_data.get('pre_script')
    if pre_script and str(pre_script).strip():
        hash_tree.append({
            "type": "JSR223PreProcessor",
            "clazzName": "io.metersphere.api.dto.definition.request.processors.pre.MsJSR223PreProcessor",
            "name": "JSR223PreProcessor",
            "script": str(pre_script),
            "scriptLanguage": "groovy",
            "enable": True,
            "resourceId": str(uuid.uuid4())
        })

    # 2. JDBC 后置 (T列) - 可能是数组
    post_jdbc = parse_json_field(row_data.get('post_jdbc'), 'post_jdbc')
    if isinstance(post_jdbc, list):
        for jdbc_item in post_jdbc:
            if isinstance(jdbc_item, dict):
                hash_tree.append({
                    "type": "JDBCPostProcessor",
                    "clazzName": "io.metersphere.api.dto.definition.request.processors.post.MsJDBCPostProcessor",
                    "name": jdbc_item.get("name", "JDBCPostProcessor"),
                    "query": jdbc_item.get("query", ""),
                    "variableNames": jdbc_item.get("variableNames", ""),
                    "dataSourceId": jdbc_item.get("dataSourceId", ""),
                    "resultVariable": jdbc_item.get("resultVariable", ""),
                    "queryTimeout": jdbc_item.get("queryTimeout", 60000),
                    "enable": True,
                    "resourceId": str(uuid.uuid4())
                })

    # 3. 后置脚本 (U列)
    post_script = row_data.get('post_script')
    if post_script and str(post_script).strip():
        hash_tree.append({
            "type": "JSR223PostProcessor",
            "clazzName": "io.metersphere.api.dto.definition.request.processors.post.MsJSR223PostProcessor",
            "name": "JSR223PostProcessor",
            "script": str(post_script),
            "scriptLanguage": "groovy",
            "enable": True,
            "resourceId": str(uuid.uuid4())
        })

    # 4. 断言 (P/Q/R列)
    assert_json = parse_json_field(row_data.get('assert_json'), 'assert_json')
    assert_regex = parse_json_field(row_data.get('assert_regex'), 'assert_regex')
    assert_duration = row_data.get('assert_duration')

    # 只有当存在任一断言时才添加 Assertions 节点
    if assert_json or assert_regex or (assert_duration and str(assert_duration).strip()):
        assertion_node = {
            "type": "Assertions",
            "clazzName": "io.metersphere.api.dto.definition.request.assertions.MsAssertions",
            "name": "Assertions",
            "enable": True,
            "resourceId": str(uuid.uuid4()),
            "jsonPath": assert_json if isinstance(assert_json, list) else [],
            "regex": assert_regex if isinstance(assert_regex, list) else []
        }

        # 响应时间断言
        if assert_duration and str(assert_duration).strip():
            try:
                duration_value = int(assert_duration)
                assertion_node["duration"] = {
                    "enable": True,
                    "type": "Duration",
                    "value": duration_value
                }
            except (ValueError, TypeError):
                pass

        hash_tree.append(assertion_node)

    return hash_tree


def build_request_json(row_data: Dict) -> Dict:
    """构建 MS request JSON 结构"""
    # 解析 headers
    headers = parse_json_field(row_data.get('headers'), 'headers')
    if not isinstance(headers, list):
        headers = []

    # 构建 body
    body_type = row_data.get('body_type', 'JSON')
    body_content = row_data.get('body', '')

    body = {
        "type": str(body_type) if body_type else "JSON",
        "raw": str(body_content) if body_content else "",
        "json": True
    }

    # 构建 hashTree
    hash_tree = build_hash_tree(row_data)

    return {
        "type": "HTTPSamplerProxy",
        "clazzName": "io.metersphere.api.dto.definition.request.sampler.MsHTTPSamplerProxy",
        "method": str(row_data.get('api_method', 'POST')),
        "path": str(row_data.get('api_path', '')),
        "protocol": "HTTP",
        "enable": True,
        "active": True,
        "refType": "CASE",
        "headers": headers,
        "body": body,
        "hashTree": hash_tree
    }


def build_case_data(row_data: Dict, project_id: str, include_id: bool = False) -> Dict:
    """构建完整的用例数据

    Args:
        row_data: Excel 行数据
        project_id: 项目 ID
        include_id: 是否包含 id 字段（更新模式需要）
    """
    # 解析 tags
    tags = row_data.get('tags', '')
    if tags and isinstance(tags, str):
        tags_list = [t.strip() for t in tags.split(',') if t.strip()]
    else:
        tags_list = []

    # 构建 request JSON 字符串
    request_obj = build_request_json(row_data)
    request_json = json.dumps(request_obj, ensure_ascii=False)

    data = {
        "name": str(row_data.get('name', '')),
        "priority": str(row_data.get('priority', 'P0')),
        "description": str(row_data.get('description', '')),
        "tags": tags_list,
        "apiDefinitionId": str(row_data.get('api_uuid', '')),
        "projectId": project_id,
        "request": request_json
    }

    # 更新模式需要 id 字段
    if include_id:
        data["id"] = str(row_data.get('case_uuid', ''))

    return data


def create_case(base_url: str, access_key: str, secret_key: str, case_data: Dict) -> Optional[Dict]:
    """调用 MS API 创建用例"""
    url = f"{base_url}/api/api/testcase/create"

    headers = {
        "accessKey": access_key,
        "signature": generate_signature(access_key, secret_key),
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(url, headers=headers, json=case_data, timeout=30)
        response.raise_for_status()
        data = response.json()

        if data.get("success"):
            return data.get("data")
        else:
            print(f"API Error: {data.get('message')}", file=sys.stderr)
            return None
    except requests.exceptions.RequestException as e:
        print(f"Request Error: {e}", file=sys.stderr)
        return None


def update_case(base_url: str, access_key: str, secret_key: str, case_data: Dict) -> Optional[Dict]:
    """调用 MS API 更新用例"""
    url = f"{base_url}/api/api/testcase/update"

    headers = {
        "accessKey": access_key,
        "signature": generate_signature(access_key, secret_key),
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(url, headers=headers, json=case_data, timeout=30)
        response.raise_for_status()
        data = response.json()

        if data.get("success"):
            return data.get("data")
        else:
            print(f"API Error: {data.get('message')}", file=sys.stderr)
            return None
    except requests.exceptions.RequestException as e:
        print(f"Request Error: {e}", file=sys.stderr)
        return None


def parse_case_ids(case_ids_str: str) -> List[str]:
    """解析用例编号参数，支持逗号分隔"""
    if not case_ids_str:
        return []
    return [cid.strip() for cid in case_ids_str.split(',') if cid.strip()]


def read_excel_row(ws, row_idx: int) -> Dict:
    """读取 Excel 一行数据到字典"""
    return {
        'api_id': ws.cell(row=row_idx, column=COL_API_ID + 1).value,
        'api_uuid': ws.cell(row=row_idx, column=COL_API_UUID + 1).value,
        'case_id': ws.cell(row=row_idx, column=COL_CASE_ID + 1).value,
        'case_uuid': ws.cell(row=row_idx, column=COL_CASE_UUID + 1).value,
        'status': ws.cell(row=row_idx, column=COL_STATUS + 1).value,
        'name': ws.cell(row=row_idx, column=COL_NAME + 1).value,
        'priority': ws.cell(row=row_idx, column=COL_PRIORITY + 1).value,
        'test_summary': ws.cell(row=row_idx, column=COL_TEST_SUMMARY + 1).value,
        'description': ws.cell(row=row_idx, column=COL_DESCRIPTION + 1).value,
        'tags': ws.cell(row=row_idx, column=COL_TAGS + 1).value,
        'api_path': ws.cell(row=row_idx, column=COL_API_PATH + 1).value,
        'api_method': ws.cell(row=row_idx, column=COL_API_METHOD + 1).value,
        'headers': ws.cell(row=row_idx, column=COL_HEADERS + 1).value,
        'body_type': ws.cell(row=row_idx, column=COL_BODY_TYPE + 1).value,
        'body': ws.cell(row=row_idx, column=COL_BODY + 1).value,
        'assert_json': ws.cell(row=row_idx, column=COL_ASSERT_JSON + 1).value,
        'assert_regex': ws.cell(row=row_idx, column=COL_ASSERT_REGEX + 1).value,
        'assert_duration': ws.cell(row=row_idx, column=COL_ASSERT_DURATION + 1).value,
        'pre_script': ws.cell(row=row_idx, column=COL_PRE_SCRIPT + 1).value,
        'post_jdbc': ws.cell(row=row_idx, column=COL_POST_JDBC + 1).value,
        'post_script': ws.cell(row=row_idx, column=COL_POST_SCRIPT + 1).value,
    }


def main():
    parser = argparse.ArgumentParser(description="Import/Update test cases from Excel to MeterSphere")
    parser.add_argument("--base-url", required=True, help="MeterSphere base URL")
    parser.add_argument("--access-key", required=True, help="API access key")
    parser.add_argument("--secret-key", required=True, help="API secret key")
    parser.add_argument("--project-id", required=True, help="MeterSphere project ID")
    parser.add_argument("--excel-path", required=True, help="Path to Excel file")
    parser.add_argument("--mode", choices=["import", "update"], default="import",
                        help="Operation mode: import (create new) or update (modify existing)")
    parser.add_argument("--case-ids", default="",
                        help="Case IDs to update (comma-separated, required for update mode)")

    args = parser.parse_args()

    is_update_mode = args.mode == "update"
    target_case_ids = []

    # 更新模式需要指定 case-ids
    if is_update_mode:
        target_case_ids = parse_case_ids(args.case_ids)
        if not target_case_ids:
            print("Error: --case-ids is required for update mode", file=sys.stderr)
            sys.exit(1)
        print(f"Update mode - Target case IDs: {target_case_ids}")
    else:
        print("Import mode - Processing cases with status='待导入'")

    # 加载 Excel
    try:
        wb = load_workbook(args.excel_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel: {e}", file=sys.stderr)
        sys.exit(1)

    # 统计
    success_count = 0
    fail_count = 0
    skip_count = 0

    # 遍历行 (从第3行开始，跳过表头和说明行)
    for row_idx in range(3, ws.max_row + 1):
        row_data = read_excel_row(ws, row_idx)

        # 根据模式筛选
        if is_update_mode:
            # 更新模式：按 case_id 筛选
            case_id = row_data.get('case_id')
            if str(case_id) not in target_case_ids:
                skip_count += 1
                continue
        else:
            # 导入模式：按状态筛选
            status = row_data.get('status')
            if status != '待导入':
                skip_count += 1
                continue

        # 检查必填字段
        name = row_data.get('name')
        api_uuid = row_data.get('api_uuid')
        if not name or not api_uuid:
            print(f"Row {row_idx}: Missing required fields (name or api_uuid)", file=sys.stderr)
            fail_count += 1
            continue

        # 更新模式额外检查 case_uuid
        if is_update_mode:
            case_uuid = row_data.get('case_uuid')
            if not case_uuid:
                print(f"Row {row_idx}: 更新需要 case_uuid（D列）", file=sys.stderr)
                fail_count += 1
                continue

        # 构建用例数据
        case_data = build_case_data(row_data, args.project_id, include_id=is_update_mode)

        # 执行导入或更新
        if is_update_mode:
            print(f"Updating case {row_data.get('case_id')}: {name}...", end=" ")
            result = update_case(args.base_url, args.access_key, args.secret_key, case_data)
            if result:
                ws.cell(row=row_idx, column=COL_STATUS + 1, value='已更新')
                print("OK")
                success_count += 1
            else:
                print("FAILED")
                fail_count += 1
        else:
            print(f"Creating case: {name}...", end=" ")
            result = create_case(args.base_url, args.access_key, args.secret_key, case_data)
            if result:
                case_id = result.get('num', '')
                case_uuid = result.get('id', '')
                ws.cell(row=row_idx, column=COL_CASE_ID + 1, value=case_id)
                ws.cell(row=row_idx, column=COL_CASE_UUID + 1, value=case_uuid)
                ws.cell(row=row_idx, column=COL_STATUS + 1, value='已导入')
                print(f"OK (ID: {case_id})")
                success_count += 1
            else:
                print("FAILED")
                fail_count += 1

    # 保存 Excel
    if success_count > 0:
        try:
            wb.save(args.excel_path)
            print(f"\nExcel saved: {args.excel_path}")
        except Exception as e:
            print(f"\nError saving Excel: {e}", file=sys.stderr)

    # 输出摘要
    mode_label = "Update" if is_update_mode else "Import"
    print(f"\n{'='*50}")
    print(f"{mode_label} Summary:")
    print(f"  Success: {success_count}")
    print(f"  Failed:  {fail_count}")
    print(f"  Skipped: {skip_count}")
    print(f"{'='*50}")

    # 返回结果 JSON
    result = {
        "mode": args.mode,
        "success": success_count,
        "failed": fail_count,
        "skipped": skip_count,
        "excel_path": args.excel_path
    }
    print(f"\nResult: {json.dumps(result, ensure_ascii=False)}")

    sys.exit(0 if fail_count == 0 else 1)


if __name__ == "__main__":
    main()
