#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MeterSphere 场景批量导入/更新
从 Excel 读取场景数据，基于模板构建场景定义，创建或更新场景并回写结果

支持两种模式：
- import: 导入新场景（筛选 status=待导入）
- update: 更新已有场景（通过 --case-nums 指定编号）
"""

import argparse
import base64
import copy
import json
import sys
import time
import uuid
from typing import Dict, List, Optional, Any

try:
    import requests
except ImportError:
    print("Error: requests module not found. Please install: pip3 install requests", file=sys.stderr)
    sys.exit(1)

try:
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.backends import default_backend
    from cryptography.hazmat.primitives import padding
except ImportError:
    print("Error: cryptography module not found. Please install: pip3 install cryptography", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl module not found. Please install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)



# Sheet1 列映射 (0-indexed)
COL_SCENARIO_ID = 0      # A: scenario_id
COL_SCENARIO_UUID = 1    # B: scenario_uuid
COL_STATUS = 2           # C: status
COL_CASE_NUM = 3         # D: case_num
COL_CATEGORY = 4         # E: category
COL_NAME = 5             # F: name
COL_PRIORITY = 6         # G: priority
COL_PRECONDITION = 7     # H: precondition
COL_TEST_STEPS = 8       # I: test_steps
COL_EXPECTED = 9         # J: expected
COL_REMARK = 10          # K: remark
COL_MODULE = 11          # L: module

# Sheet2 列映射 (0-indexed)
S2_CASE_NUM = 0          # A: case_num
S2_STEP_INDEX = 1        # B: step_index
S2_STEP_TYPE = 2         # C: step_type
S2_STEP_NAME = 3         # D: step_name
S2_PATH = 4              # E: path
S2_METHOD = 5            # F: method
S2_HEADERS = 6           # G: headers
S2_BODY = 7              # H: body
S2_BODY_TYPE = 8         # I: body_type
S2_ASSERTIONS = 9        # J: assertions
S2_EXTRACT_VARS = 10     # K: extract_vars
S2_PRE_SCRIPT = 11       # L: pre_script
S2_POST_SCRIPT = 12      # M: post_script
S2_JDBC_QUERY = 13       # N: jdbc_query
S2_JDBC_VARIABLES = 14   # O: jdbc_variables


def generate_signature(access_key: str, secret_key: str) -> str:
    """生成 API 签名（AES-128-CBC）"""
    timestamp = str(int(time.time() * 1000))
    plain_text = f"{access_key}|{timestamp}"

    key = secret_key.encode('utf-8')[:16].ljust(16, b'\x00')
    iv = access_key.encode('utf-8')[:16].ljust(16, b'\x00')

    padder = padding.PKCS7(128).padder()
    padded_data = padder.update(plain_text.encode('utf-8')) + padder.finalize()

    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted = encryptor.update(padded_data) + encryptor.finalize()

    return base64.b64encode(encrypted).decode('utf-8')


def parse_json_field(value: Any, field_name: str) -> Any:
    """解析 JSON 字段，返回解析后的对象或空列表"""
    if not value:
        return []
    if isinstance(value, (list, dict)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError as e:
        print(f"Warning: Failed to parse {field_name}: {e}", file=sys.stderr)
        return []


def export_template_scenario(base_url: str, access_key: str, secret_key: str,
                             scenario_id: str) -> Optional[Dict]:
    """导出模板场景的完整定义"""
    url = f"{base_url}/api/api/automation/export"

    headers = {
        "accessKey": access_key,
        "signature": generate_signature(access_key, secret_key),
        "Content-Type": "application/json"
    }

    payload = {"ids": [scenario_id]}

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        data = response.json()

        if data.get("success") and data.get("data"):
            scenarios = data["data"]
            if scenarios and len(scenarios) > 0:
                return scenarios[0]
        print(f"Export Error: {data.get('message', 'No data returned')}", file=sys.stderr)
        return None
    except requests.exceptions.RequestException as e:
        print(f"Export Request Error: {e}", file=sys.stderr)
        return None


def fix_step_references(step: Dict) -> Dict:
    """修复克隆步骤中的引用问题"""
    if 'referenced' in step:
        step['referenced'] = 'REF'

    step['resourceId'] = str(uuid.uuid4())
    if 'id' in step:
        step['id'] = str(uuid.uuid4())

    if 'hashTree' in step and step['hashTree']:
        for sub in step['hashTree']:
            fix_step_references(sub)

    return step


def build_step_from_sheet2(step_data: Dict, template_step: Dict) -> Dict:
    """根据 Sheet2 数据修改模板步骤"""
    new_step = copy.deepcopy(template_step)
    fix_step_references(new_step)

    # 更新 body
    if step_data.get('body'):
        if 'body' not in new_step:
            new_step['body'] = {}
        new_step['body']['raw'] = str(step_data['body'])

    # 更新 headers
    if step_data.get('headers'):
        headers = parse_json_field(step_data['headers'], 'headers')
        if headers:
            new_step['headers'] = headers

    # 更新断言
    if step_data.get('assertions'):
        assertions = parse_json_field(step_data['assertions'], 'assertions')
        if assertions and 'hashTree' in new_step:
            for sub in new_step['hashTree']:
                if sub.get('type') == 'Assertions':
                    sub['jsonPath'] = assertions
                    break

    # 更新变量提取
    if step_data.get('extract_vars'):
        extract_vars = parse_json_field(step_data['extract_vars'], 'extract_vars')
        if extract_vars and 'hashTree' in new_step:
            for sub in new_step['hashTree']:
                if sub.get('type') == 'Extract':
                    sub['json'] = extract_vars
                    break

    # 更新前置脚本
    if step_data.get('pre_script') and 'hashTree' in new_step:
        for sub in new_step['hashTree']:
            if sub.get('type') == 'JSR223PreProcessor':
                sub['script'] = str(step_data['pre_script'])
                break

    # 更新后置脚本
    if step_data.get('post_script') and 'hashTree' in new_step:
        for sub in new_step['hashTree']:
            if sub.get('type') == 'JSR223PostProcessor':
                sub['script'] = str(step_data['post_script'])
                break

    return new_step


def build_scenario_definition(template_def: Dict, steps_data: List[Dict]) -> Dict:
    """基于模板和 Sheet2 数据构建场景定义"""
    new_def = copy.deepcopy(template_def)
    template_steps = new_def.get('hashTree', [])

    if not steps_data:
        # 没有步骤数据，保留模板全部步骤但修复引用
        for step in template_steps:
            fix_step_references(step)
        return new_def

    # 根据 Sheet2 步骤数据构建新的 hashTree
    new_hash_tree = []
    for step_data in steps_data:
        step_index = int(step_data.get('step_index', 1)) - 1  # 转为 0-indexed
        if step_index < len(template_steps):
            new_step = build_step_from_sheet2(step_data, template_steps[step_index])
            new_hash_tree.append(new_step)

    new_def['hashTree'] = new_hash_tree
    return new_def


def build_scenario_data(row_data: Dict, steps_data: List[Dict], template: Dict,
                        project_id: str, module_id: str, include_id: bool = False) -> Dict:
    """构建完整的场景数据

    Args:
        row_data: Excel 行数据
        steps_data: Sheet2 步骤数据
        template: 模板场景
        project_id: 项目 ID
        module_id: 模块 ID
        include_id: 是否包含 id 字段（更新模式需要）
    """

    # 构建场景定义
    template_def = template.get('scenarioDefinition', {})
    scenario_def = build_scenario_definition(template_def, steps_data)

    data = {
        "projectId": project_id,
        "apiScenarioModuleId": module_id,
        "name": str(row_data.get('name', '')),
        "level": str(row_data.get('priority', 'P0')),
        "status": "Underway",
        "principal": template.get('principal', ''),
        "stepTotal": len(steps_data) if steps_data else len(template_def.get('hashTree', [])),
        "description": str(row_data.get('remark', '')),
        "scenarioDefinition": scenario_def
    }

    # 更新模式需要 id 字段
    if include_id:
        data["id"] = str(row_data.get('scenario_uuid', ''))

    return data


def add_resource_ids(obj: Any) -> Any:
    """递归为所有节点添加 resourceId"""
    if isinstance(obj, dict):
        if "type" in obj and "resourceId" not in obj:
            obj["resourceId"] = str(uuid.uuid4())
        for key, value in obj.items():
            obj[key] = add_resource_ids(value)
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            obj[i] = add_resource_ids(item)
    return obj


def create_scenario(base_url: str, access_key: str, secret_key: str,
                    scenario_data: Dict) -> Optional[Dict]:
    """调用 MS API 创建场景"""
    url = f"{base_url}/api/api/automation/create"

    headers = {
        "accessKey": access_key,
        "signature": generate_signature(access_key, secret_key),
    }

    scenario_data = add_resource_ids(scenario_data)
    json_str = json.dumps(scenario_data, ensure_ascii=False)
    files = {
        'request': (None, json_str, 'application/json')
    }

    try:
        response = requests.post(url, headers=headers, files=files, timeout=60)

        if response.status_code == 200:
            data = response.json()
            if data.get("success"):
                result = data.get("data", {})
                return result if isinstance(result, dict) else {"id": result}
            else:
                print(f"API Error: {data.get('message')}", file=sys.stderr)
                return None
        else:
            print(f"HTTP {response.status_code}: {response.text[:200]}", file=sys.stderr)
            return None
    except requests.exceptions.RequestException as e:
        print(f"Request Error: {e}", file=sys.stderr)
        return None


def update_scenario(base_url: str, access_key: str, secret_key: str,
                    scenario_data: Dict) -> Optional[Dict]:
    """调用 MS API 更新场景"""
    url = f"{base_url}/api/api/automation/update"

    headers = {
        "accessKey": access_key,
        "signature": generate_signature(access_key, secret_key),
    }

    scenario_data = add_resource_ids(scenario_data)
    json_str = json.dumps(scenario_data, ensure_ascii=False)
    files = {
        'request': (None, json_str, 'application/json')
    }

    try:
        response = requests.post(url, headers=headers, files=files, timeout=60)

        if response.status_code == 200:
            data = response.json()
            if data.get("success"):
                result = data.get("data", {})
                return result if isinstance(result, dict) else {"id": result}
            else:
                print(f"API Error: {data.get('message')}", file=sys.stderr)
                return None
        else:
            print(f"HTTP {response.status_code}: {response.text[:200]}", file=sys.stderr)
            return None
    except requests.exceptions.RequestException as e:
        print(f"Request Error: {e}", file=sys.stderr)
        return None


def parse_case_nums(case_nums_str: str) -> List[str]:
    """解析场景编号参数，支持逗号分隔"""
    if not case_nums_str:
        return []
    return [cn.strip() for cn in case_nums_str.split(',') if cn.strip()]


def read_sheet1_row(ws, row_idx: int) -> Dict:
    """读取 Sheet1 一行数据"""
    return {
        'scenario_id': ws.cell(row=row_idx, column=COL_SCENARIO_ID + 1).value,
        'scenario_uuid': ws.cell(row=row_idx, column=COL_SCENARIO_UUID + 1).value,
        'status': ws.cell(row=row_idx, column=COL_STATUS + 1).value,
        'case_num': ws.cell(row=row_idx, column=COL_CASE_NUM + 1).value,
        'category': ws.cell(row=row_idx, column=COL_CATEGORY + 1).value,
        'name': ws.cell(row=row_idx, column=COL_NAME + 1).value,
        'priority': ws.cell(row=row_idx, column=COL_PRIORITY + 1).value,
        'precondition': ws.cell(row=row_idx, column=COL_PRECONDITION + 1).value,
        'test_steps': ws.cell(row=row_idx, column=COL_TEST_STEPS + 1).value,
        'expected': ws.cell(row=row_idx, column=COL_EXPECTED + 1).value,
        'remark': ws.cell(row=row_idx, column=COL_REMARK + 1).value,
        'module': ws.cell(row=row_idx, column=COL_MODULE + 1).value,
    }


def read_sheet2_steps(ws, case_num: str) -> List[Dict]:
    """读取 Sheet2 中指定 case_num 的所有步骤"""
    steps = []
    for row_idx in range(2, ws.max_row + 1):  # 从第2行开始，跳过表头
        row_case_num = ws.cell(row=row_idx, column=S2_CASE_NUM + 1).value
        if str(row_case_num) == str(case_num):
            steps.append({
                'case_num': row_case_num,
                'step_index': ws.cell(row=row_idx, column=S2_STEP_INDEX + 1).value,
                'step_type': ws.cell(row=row_idx, column=S2_STEP_TYPE + 1).value,
                'step_name': ws.cell(row=row_idx, column=S2_STEP_NAME + 1).value,
                'path': ws.cell(row=row_idx, column=S2_PATH + 1).value,
                'method': ws.cell(row=row_idx, column=S2_METHOD + 1).value,
                'headers': ws.cell(row=row_idx, column=S2_HEADERS + 1).value,
                'body': ws.cell(row=row_idx, column=S2_BODY + 1).value,
                'body_type': ws.cell(row=row_idx, column=S2_BODY_TYPE + 1).value,
                'assertions': ws.cell(row=row_idx, column=S2_ASSERTIONS + 1).value,
                'extract_vars': ws.cell(row=row_idx, column=S2_EXTRACT_VARS + 1).value,
                'pre_script': ws.cell(row=row_idx, column=S2_PRE_SCRIPT + 1).value,
                'post_script': ws.cell(row=row_idx, column=S2_POST_SCRIPT + 1).value,
                'jdbc_query': ws.cell(row=row_idx, column=S2_JDBC_QUERY + 1).value,
                'jdbc_variables': ws.cell(row=row_idx, column=S2_JDBC_VARIABLES + 1).value,
            })
    return sorted(steps, key=lambda x: int(x.get('step_index', 0)))


def main():
    parser = argparse.ArgumentParser(description="Import/Update scenarios from Excel to MeterSphere")
    parser.add_argument("--base-url", required=True, help="MeterSphere base URL")
    parser.add_argument("--access-key", required=True, help="API access key")
    parser.add_argument("--secret-key", required=True, help="API secret key")
    parser.add_argument("--project-id", required=True, help="MeterSphere project ID")
    parser.add_argument("--template-id", required=True, help="Template scenario ID")
    parser.add_argument("--excel-path", required=True, help="Path to Excel file")
    parser.add_argument("--mode", choices=["import", "update"], default="import",
                        help="Operation mode: import (create new) or update (modify existing)")
    parser.add_argument("--case-nums", default="",
                        help="Case numbers to update (comma-separated, required for update mode)")
    parser.add_argument("--module-id", required=True, help="Module ID for scenarios")

    args = parser.parse_args()

    is_update_mode = args.mode == "update"
    target_case_nums = []

    # 更新模式需要指定 case-nums
    if is_update_mode:
        target_case_nums = parse_case_nums(args.case_nums)
        if not target_case_nums:
            print("Error: --case-nums is required for update mode", file=sys.stderr)
            sys.exit(1)
        print(f"Update mode - Target case numbers: {target_case_nums}")
    else:
        print("Import mode - Processing scenarios with status='待导入'")

    # 加载 Excel
    try:
        wb = load_workbook(args.excel_path)
        ws1 = wb.worksheets[0]  # Sheet1: 场景用例
        ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else None  # Sheet2: 步骤详情
    except Exception as e:
        print(f"Error loading Excel: {e}", file=sys.stderr)
        sys.exit(1)

    # 导出模板场景
    print(f"Exporting template scenario: {args.template_id}...")
    template = export_template_scenario(args.base_url, args.access_key, args.secret_key, args.template_id)
    if not template:
        print("Failed to export template scenario", file=sys.stderr)
        sys.exit(1)
    print(f"Template loaded: {template.get('name')}")

    # 统计
    success_count = 0
    fail_count = 0
    skip_count = 0

    # 遍历 Sheet1 (从第2行开始，跳过表头)
    for row_idx in range(2, ws1.max_row + 1):
        row_data = read_sheet1_row(ws1, row_idx)

        # 根据模式筛选
        if is_update_mode:
            # 更新模式：按 case_num 筛选
            case_num = row_data.get('case_num')
            if str(case_num) not in target_case_nums:
                skip_count += 1
                continue
        else:
            # 导入模式：按状态筛选
            status = row_data.get('status')
            if status != '待导入':
                skip_count += 1
                continue

        # 检查必填字段
        name = row_data.get('name')
        case_num = row_data.get('case_num')
        if not name:
            print(f"Row {row_idx}: Missing required field (name)", file=sys.stderr)
            fail_count += 1
            continue

        # 更新模式额外检查 scenario_uuid
        if is_update_mode:
            scenario_uuid = row_data.get('scenario_uuid')
            if not scenario_uuid:
                print(f"Row {row_idx}: 更新需要 scenario_uuid（B列）", file=sys.stderr)
                fail_count += 1
                continue

        # 读取 Sheet2 步骤数据
        steps_data = []
        if ws2 and case_num:
            steps_data = read_sheet2_steps(ws2, case_num)

        # 构建场景数据
        scenario_data = build_scenario_data(
            row_data, steps_data, template,
            args.project_id, args.module_id, include_id=is_update_mode
        )

        # 执行导入或更新
        if is_update_mode:
            print(f"Updating scenario {case_num}: {name}...", end=" ")
            result = update_scenario(args.base_url, args.access_key, args.secret_key, scenario_data)
            if result:
                ws1.cell(row=row_idx, column=COL_STATUS + 1, value='已更新')
                print("OK")
                success_count += 1
            else:
                print("FAILED")
                fail_count += 1
        else:
            print(f"Creating scenario: {name}...", end=" ")
            result = create_scenario(args.base_url, args.access_key, args.secret_key, scenario_data)
            if result:
                scenario_id = result.get('num', '')
                scenario_uuid = result.get('id', '')
                ws1.cell(row=row_idx, column=COL_SCENARIO_ID + 1, value=scenario_id)
                ws1.cell(row=row_idx, column=COL_SCENARIO_UUID + 1, value=scenario_uuid)
                ws1.cell(row=row_idx, column=COL_STATUS + 1, value='已导入')
                print(f"OK (ID: {scenario_id})")
                success_count += 1
            else:
                print("FAILED")
                fail_count += 1

    # 保存 Excel
    if success_count > 0:
        try:
            wb.save(args.excel_path)
            print(f"\nExcel saved: {args.excel_path}")
        except Exception as e:
            print(f"\nError saving Excel: {e}", file=sys.stderr)

    # 输出摘要
    mode_label = "Update" if is_update_mode else "Import"
    print(f"\n{'='*50}")
    print(f"{mode_label} Summary:")
    print(f"  Success: {success_count}")
    print(f"  Failed:  {fail_count}")
    print(f"  Skipped: {skip_count}")
    print(f"{'='*50}")

    # 返回结果 JSON
    result = {
        "mode": args.mode,
        "success": success_count,
        "failed": fail_count,
        "skipped": skip_count,
        "excel_path": args.excel_path
    }
    print(f"\nResult: {json.dumps(result, ensure_ascii=False)}")

    sys.exit(0 if fail_count == 0 else 1)


if __name__ == "__main__":
    main()
